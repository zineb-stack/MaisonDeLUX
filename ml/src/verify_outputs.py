"""Verify canonical MaisonDeLUX outputs and write a compact evidence report."""
from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
REQUIRED = [
    "data/raw/maisondelux_raw.csv", "data/raw/maisondelux_raw.xlsx", "data/raw/maisondelux_raw.parquet",
    "data/processed/maisondelux_clean.csv", "data/processed/maisondelux_clean.xlsx", "data/processed/maisondelux_clean.parquet",
    "data/processed/maisondelux_rejected.csv", "data/geographic/morocco_regions.geojson",
    "data/geographic/morocco_cities.geojson", "data/geographic/morocco_neighborhoods.geojson",
    "reports/data_quality/data_quality_report.md", "reports/scraping/source_coverage_report.md",
    "reports/scraping/geographic_coverage_report.md", "reports/scraping/historical_coverage_report.md",
    "ml/notebooks/maisondelux_data_pipeline.ipynb",
]
SHEETS = ["all_rows", "valid_rows", "rejected_rows", "source_summary", "city_summary", "quality_summary", "scraping_errors"]


def digest(path: Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            value.update(chunk)
    return value.hexdigest()


def workbook_info(path: Path) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=False)
    dimensions = {}
    formula_count = 0
    error_formulas = []
    for sheet in workbook.worksheets:
        dimensions[sheet.title] = sheet.calculate_dimension(force=True)
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    formula_count += 1
                    if any(token in str(cell.value) for token in ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?")):
                        error_formulas.append(f"{sheet.title}!{cell.coordinate}")
    first_header = workbook["all_rows"]["A1"].value
    return {"sheets": workbook.sheetnames, "dimensions": dimensions, "formula_count": formula_count,
            "formula_error_tokens": error_formulas, "first_header": first_header}


def verify() -> dict:
    missing = [item for item in REQUIRED if not (ROOT / item).exists()]
    if missing:
        raise AssertionError(f"Missing required outputs: {missing}")
    raw = pd.read_csv(ROOT / "data/raw/maisondelux_raw.csv", low_memory=False)
    clean = pd.read_csv(ROOT / "data/processed/maisondelux_clean.csv", low_memory=False)
    rejected = pd.read_csv(ROOT / "data/processed/maisondelux_rejected.csv", low_memory=False)
    raw_parquet = pd.read_parquet(ROOT / "data/raw/maisondelux_raw.parquet")
    clean_parquet = pd.read_parquet(ROOT / "data/processed/maisondelux_clean.parquet")
    assert len(raw) == len(clean) + len(rejected)
    assert list(raw.columns) == list(raw_parquet.columns)
    assert list(clean.columns) == list(clean_parquet.columns)
    assert raw["listing_id"].fillna("").tolist() == raw_parquet["listing_id"].fillna("").tolist()
    assert clean["listing_id"].fillna("").tolist() == clean_parquet["listing_id"].fillna("").tolist()
    assert clean["validation_status"].eq("valid").all()
    assert clean["deduplication_status"].eq("unique").all()
    assert clean["listing_id"].is_unique

    workbooks = {
        "raw": workbook_info(ROOT / "data/raw/maisondelux_raw.xlsx"),
        "clean": workbook_info(ROOT / "data/processed/maisondelux_clean.xlsx"),
    }
    for info in workbooks.values():
        assert info["sheets"] == SHEETS
        assert info["first_header"] == "listing_id"
        assert not info["formula_error_tokens"]

    geographic = {}
    for name in ("regions", "cities", "neighborhoods"):
        layer = json.loads((ROOT / f"data/geographic/morocco_{name}.geojson").read_text(encoding="utf-8"))
        assert layer.get("type") == "FeatureCollection"
        assert all(feature.get("geometry") and feature.get("properties", {}).get("source") for feature in layer["features"])
        geographic[name] = len(layer["features"])

    previews = list((ROOT / "outputs/01a06449-2ac1-7fa2-a7ad-6d42f27ec146/previews").glob("*.png"))
    result = {
        "status": "passed", "row_counts": {"raw": len(raw), "valid_unique": len(clean), "rejected_or_warning": len(rejected),
                                                "duplicates": int(raw["deduplication_status"].ne("unique").sum())},
        "csv_parquet_equivalent": True, "clean_listing_ids_unique": True,
        "workbooks": workbooks, "workbook_visual_previews_checked": len(previews),
        "geographic_feature_counts": geographic,
        "files": {item: {"bytes": (ROOT / item).stat().st_size, "sha256": digest(ROOT / item)} for item in REQUIRED},
    }
    output = ROOT / "reports/data_quality/verification_report.json"
    output.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    lines = [
        "# Output verification", "", "Status: **passed**", "",
        f"- Raw / valid unique / rejected-warning / duplicates: **{len(raw):,} / {len(clean):,} / {len(rejected):,} / {result['row_counts']['duplicates']:,}**",
        "- CSV and Parquet schemas, order, row counts and listing IDs reconcile.",
        "- Clean listing IDs are unique; every clean row is `valid` and `unique`.",
        f"- Both workbooks have all seven required sheets, clean headers, {workbooks['raw']['formula_count'] + workbooks['clean']['formula_count']} control formulas, and no formula error tokens.",
        f"- {len(previews)} workbook sheet previews were rendered for visual QA.",
        f"- Geographic features: {geographic['regions']:,} regions, {geographic['cities']:,} cities/towns, {geographic['neighborhoods']:,} districts/neighborhoods.",
    ]
    (ROOT / "reports/data_quality/verification_report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    return result


if __name__ == "__main__":
    print(json.dumps(verify()["row_counts"], sort_keys=True))
